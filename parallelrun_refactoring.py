import pandas as pd
import os
import time
from multiprocessing import Pool
import multiprocessing
from openpyxl import load_workbook
import re
import sys
from pathlib import Path

# ----------------------
# Predefined functions and setup (moved from notebook)
# ----------------------
sys.path.append('./Helpers')
from collections import defaultdict
from Helpers import preprocessing, divide_adl, postprocessing, helper, auxiluary
from preprocessing import preprocess, preprocess_with_adl, get_extended_paths_with_connector_info, identify_liveness_assert, load_adl
from postprocessing import replace_attachments_in_adl, ensure_parameters_correct_output_roles, extract_fix_undefined_component_port, reorder_input_roles_first
from auxiluary import split_into_two_roles, detect_output_role_issues, extract_attach_statements, extract_assert_statements, parse_assert_components, match_asserts_to_paths
from divide_adl import get_verification_results_with_adl


# ----------------------
# Function to execute the refactoring process
# ----------------------
def build_execution_env(adl_file_name, new_requirement, run_id):
    return {
        "adl_file_name": adl_file_name,
        "new_requirement": new_requirement,
        "run_id": run_id,
        
        # Core helpers
        "split_into_two_roles": split_into_two_roles,
        "detect_output_role_issues": detect_output_role_issues,
        "extract_attach_statements": extract_attach_statements,
        "extract_assert_statements": extract_assert_statements,
        "parse_assert_components": parse_assert_components,
        "match_asserts_to_paths": match_asserts_to_paths,

        # Module imports
        "load_adl": load_adl,
        "preprocess": preprocess,
        "preprocess_with_adl": preprocess_with_adl,
        "get_extended_paths_with_connector_info": get_extended_paths_with_connector_info,
        "identify_liveness_assert": identify_liveness_assert,
        "replace_attachments_in_adl": replace_attachments_in_adl,
        "ensure_parameters_correct_output_roles": ensure_parameters_correct_output_roles,
        "extract_fix_undefined_component_port": extract_fix_undefined_component_port,
        "reorder_input_roles_first": reorder_input_roles_first,

        # Standard modules
        "sys": sys,
        "pd": pd,
        "Path": Path,
        "defaultdict": defaultdict,
    }



def run_one_sample(args):
    import gc
    import os
    
    print("Running one sample:", args)
    adl_file_name, new_requirement, run_id = args
    start_time = time.time()

    try:
        # Force garbage collection before starting
        gc.collect()
        
        # Read the content of refactoring.py
        with open('./Helpers/refactoring.py', 'r') as file:
            global_script = file.read()
        
        # Create a namespace dictionary with required variables
        namespace = build_execution_env(adl_file_name, new_requirement, run_id)
        
        # Execute the global script
        exec(global_script, namespace)
        
        # Get the results from the namespace
        result = namespace.get('Final_VV_result', 'No result produced')
        current_ADL = namespace.get('current_ADL', None)
        
        elapsed = time.time() - start_time
        
        # Validate return value size to prevent pipe errors
        # If ADL is too large, truncate it (multiprocessing has limits)
        if current_ADL and len(str(current_ADL)) > 10 * 1024 * 1024:  # 10MB limit
            print(f"Warning: ADL result is very large ({len(str(current_ADL))} bytes), truncating...")
            current_ADL = str(current_ADL)[:10 * 1024 * 1024] + "\n... [truncated]"
        
        # Clean up namespace to prevent memory leaks
        namespace.clear()
        gc.collect()
        
        return [adl_file_name, new_requirement, run_id, result, current_ADL, elapsed]
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        elapsed = time.time() - start_time
        print(f"Error in run_one_sample: {error_details}")
        
        # Force cleanup on error
        try:
            gc.collect()
        except:
            pass
        
        # Return a simple, serializable result
        error_msg = f"Execution Error: {str(e)}"
        # Truncate error message if too long
        if len(error_msg) > 1000:
            error_msg = error_msg[:1000] + "... [truncated]"
        
        return [adl_file_name, new_requirement, run_id, error_msg, None, elapsed]
    
def get_system_execution_paths_differences(original_paths, new_paths):
    """
    Calculate the differences between original and new system execution paths.
    
    Args:
        original_paths (list): List of original paths, where each path is a list of steps
        new_paths (list): List of new paths, where each path is a list of steps
    
    Returns:
        int: Number of paths that were added and removed
    """
    # Convert paths to tuples for hashability and comparison
    original_paths_set = {tuple(path) for path in original_paths}
    new_paths_set = {tuple(path) for path in new_paths}
    
    # Find paths that were added and removed
    added_paths = new_paths_set - original_paths_set
    removed_paths = original_paths_set - new_paths_set
    
    # Convert back to lists for the return value
    added_paths_list = [list(path) for path in added_paths]
    removed_paths_list = [list(path) for path in removed_paths]
    
    return len(added_paths_list)+ len(removed_paths_list)

# ----------------------
# Main function
# ----------------------
def main():
    print("Starting main()")
    input_file = "Refactoring_data.xlsx"
    output_file = "Refactoring_Execution_results.xlsx"
    output_file_selection = "Final_Refactoring_Execution_results.xlsx"
    start_index = 0
    end_index = 60
    
    df = pd.read_excel(input_file, engine="openpyxl")
    df = df.iloc[start_index:end_index]

    # Prepare Excel file with headers if not exist
    if not os.path.exists(output_file):
        df_empty = pd.DataFrame(columns=["ADL File Name", "New Requirement", "Run ID", "Execution Result", "Runtime (s)"])
        df_empty.to_excel(output_file, index=False, engine="openpyxl")
    if not os.path.exists(output_file_selection):
        df_empty = pd.DataFrame(columns=["ADL File Name", "New Requirement", "Run ID", "Execution Result", "Runtime (s)"])
        df_empty.to_excel(output_file_selection, index=False, engine="openpyxl")

    for idx, row in df.iterrows():
        adl_file_name = row["System"]
        new_requirement = row["Requirement"]

        # Create argument list for parallel runs
        num_runs = 7  # Number of parallel runs (change this to 1, 2, 3, etc.)
        args_list = [
            (adl_file_name, new_requirement, run_id)
            for run_id in range(1, num_runs + 1)
        ]

        # Use context manager and explicit cleanup for better memory management
        results = []
        try:
            with Pool(processes=num_runs, maxtasksperchild=1) as pool:  # Processes match num_runs
                # Use map_async with timeout to prevent hanging
                async_result = pool.map_async(run_one_sample, args_list)
                try:
                    # Wait with timeout (1 hour max per batch)
                    results = async_result.get(timeout=3600)
                except TimeoutError:
                    print(f"Timeout waiting for results from {adl_file_name}")
                    pool.terminate()
                    results = [[adl_file_name, new_requirement, run_id, "Timeout Error", None, 0.0] 
                              for run_id in range(1, num_runs + 1)]
                except Exception as e:
                    print(f"Error getting results from pool: {e}")
                    pool.terminate()
                    results = [[adl_file_name, new_requirement, run_id, f"Pool Error: {str(e)}", None, 0.0] 
                              for run_id in range(1, num_runs + 1)]
        except Exception as e:
            print(f"Error creating pool: {e}")
            import traceback
            traceback.print_exc()
            # Fallback: run sequentially if pool fails
            results = [run_one_sample(args) for args in args_list]
        
        # Force garbage collection after each batch
        import gc
        gc.collect()

        original_path = preprocess_with_adl(load_adl(adl_file_name))
        filtered_results = []  # Store results that pass validation and verification
        filtered_path_changes = []  # Store path changes for filtered results

        for result in results:
            current_adl = result[4]  # This is current_ADL
            if current_adl:
                # get verification and validation results
                from Helpers.querygrag import validation_verification
                exec_result = validation_verification(current_adl, new_requirement)
                trimmed_result = [result[0], result[1], result[2], exec_result, result[5]]

                # get path changes
                new_path = preprocess_with_adl(current_adl)
                path_change = get_system_execution_paths_differences(original_path, new_path)

                # VV filtering: Check if result passes both validation and verification
                if isinstance(exec_result, str) and exec_result.startswith("Both Validation and Verification:"):
                    filtered_results.append(trimmed_result)
                    filtered_path_changes.append(path_change)
            else:
                trimmed_result = [result[0], result[1], result[2], result[3], result[5]]

            # Write all results to output_file
            result_df = pd.DataFrame([trimmed_result], columns=["ADL File Name", "New Requirement", "Run ID", "Execution Result", "Runtime (s)"])
            append_to_excel(output_file, result_df)

        # Selection mechanism: select the result from filtered results
        if filtered_results: 
            min_path_change_idx = filtered_path_changes.index(min(filtered_path_changes))
            optimal_result = filtered_results[min_path_change_idx]
            optimal_df = pd.DataFrame([optimal_result], columns=["ADL File Name", "New Requirement", "Run ID", "Execution Result", "Runtime (s)"])
            append_to_excel(output_file_selection, optimal_df)

# ----------------------
# Append function for Excel
# ----------------------
def append_to_excel(filename, df):
    if os.path.exists(filename):
        with pd.ExcelWriter(filename, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            book = load_workbook(filename)
            sheet = writer.sheets['Sheet1']
            start_row = sheet.max_row
            df.to_excel(writer, index=False, header=False, startrow=start_row)
    else:
        df.to_excel(filename, index=False, engine="openpyxl")

# ----------------------
# Entry point
# ----------------------
if __name__ == "__main__":
    # Use spawn instead of fork to avoid memory corruption
    multiprocessing.set_start_method("spawn")
    main()
